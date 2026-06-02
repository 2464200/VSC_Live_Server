import openpyxl
import json
from collections import Counter
from pathlib import Path
import os

# Trova il file Excel più recente nella cartella Excel
excel_dir = Path(r'C:\VSC_Live_Server\Excel')
excel_files = sorted([f for f in excel_dir.iterdir() if f.suffix == '.xlsm' and not f.name.startswith('~$')], key=os.path.getmtime, reverse=True)

if not excel_files:
    print("Errore: Nessun file Excel trovato")
    exit(1)

excel_path = excel_files[0]
output_path = Path(r'C:\VSC_Live_Server\public\border_data.json')

# Log path
log_path = Path(r'C:\VSC_Live_Server\logs\report_update.log')
log_path.parent.mkdir(parents=True, exist_ok=True)

def write_log(msg):
    timestamp = __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{timestamp}] {msg}\n"
    try:
        with open(log_path, 'a', encoding='utf-8') as lf:
            lf.write(line)
    except Exception:
        pass

# Carica il workbook
try:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    write_log(f"Aperto workbook: {excel_path}")
except Exception as e:
    write_log(f"ERRORE apertura workbook: {e}")
    raise

# ============= DATI DAL FOGLIO "borderò" =============

# Foglio borderò
bordero_sheet = wb['borderò']

# 1. Totale coreografie eseguite: conta le righe attive in D12:D612
#    (qui la colonna D è il marker di presenza dati)
total_eseguite = 0
for row in range(12, 613):  # 12 a 612
    cell_value = bordero_sheet[f'D{row}'].value
    if cell_value is not None and str(cell_value).strip() != '':
        total_eseguite += 1

# 2. Coreografie eseguite con valore in colonna G: conta righe con D non vuota e G != 0
eseguite_con_valore = 0
for row in range(12, 613):
    cell_d = bordero_sheet[f'D{row}'].value
    cell_g = bordero_sheet[f'G{row}'].value
    if cell_d is not None and str(cell_d).strip() != '' and cell_g and cell_g != 0:
        eseguite_con_valore += 1

# Non eseguite: conta righe con D vuota e G != 0
non_eseguite = 0
for row in range(12, 613):
    cell_d = bordero_sheet[f'D{row}'].value
    cell_g = bordero_sheet[f'G{row}'].value
    if (cell_d is None or str(cell_d).strip() == '') and cell_g and cell_g != 0:
        non_eseguite += 1

# ============= DATI DAL FOGLIO "Accoda 8+12" =============

accoda_sheet = wb['Accoda 8+12']

# 3. Richieste ricevute con QR: conta righe non vuote in D:AA, righe 2-1000
richieste_qr = 0
for row in range(2, 1001):  # 2 a 1000
    # Controlla se almeno una cella in D:AA contiene un valore
    has_value = False
    for col in range(4, 27):  # D=4, AA=27
        cell_value = accoda_sheet.cell(row=row, column=col).value
        if cell_value:
            has_value = True
            break
    if has_value:
        richieste_qr += 1

# 4. Top 5 Coreografie: conta occorrenze di nomi in D2:AA1000
all_names = []
for row in range(2, 1001):  # 2 a 1000
    for col in range(4, 27):  # D=4, AA=27
        cell_value = accoda_sheet.cell(row=row, column=col).value
        if cell_value:
            name = str(cell_value).strip()
            if name != "XXX":  # Escludi la coreografia XXX
                all_names.append(name)

# Conta le occorrenze
name_counts = Counter(all_names)
top_5 = name_counts.most_common(5)

# 5. Coreografia più richiesta e conteggio
if top_5:
    coreo_piu_richiesta = top_5[0][0]
    conteggio_piu_richiesta = top_5[0][1]
else:
    coreo_piu_richiesta = "N/A"
    conteggio_piu_richiesta = 0

# Prepara i dati per il JSON
output_data = {
    "totale_eseguite": total_eseguite,
    "eseguite": eseguite_con_valore,
    "non_eseguite": non_eseguite,
    "richieste_qr": richieste_qr,
    "coreografia_piu_richiesta": coreo_piu_richiesta,
    "conteggio_piu_richiesta": conteggio_piu_richiesta,
    "top_5": [
        {"nome": nome, "conteggio": conteggio}
        for nome, conteggio in top_5
    ]
}

# Salva il JSON
output_path.parent.mkdir(parents=True, exist_ok=True)
try:
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    write_log(f"✓ Dati generati correttamente in {output_path}")
    write_log(f"  Totale eseguite: {output_data['totale_eseguite']}")
    write_log(f"  Eseguite (con valore in G): {output_data['eseguite']}")
    write_log(f"  Richieste QR: {output_data['richieste_qr']}")
    write_log(f"  Top coreografia: {output_data['coreografia_piu_richiesta']} ({output_data['conteggio_piu_richiesta']} volte)")
    write_log(f"  Top 5: {[item['nome'] for item in output_data['top_5']]}")
except Exception as e:
    write_log(f"ERRORE scrittura JSON: {e}")
    raise

wb.close()

write_log('Chiusura workbook')
