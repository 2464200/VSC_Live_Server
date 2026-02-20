#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script per aggiornare automaticamente border_data.json
Legge i dati dal file Excel Borderò e genera il JSON con le statistiche
"""

import openpyxl
import glob
import json
import os

def update_border_stats():
    # Leggi il file Excel
    files = glob.glob(r'c:\VSC_Live_Server\Excel\*.xlsm')
    files = [f for f in files if 'ver 13.1.46' in f and '~' not in f]
    
    if not files:
        print('Errore: File Excel non trovato')
        return False
    
    excel_file = files[0]
    print(f'Lettura file: {excel_file}')
    
    # Carica con data_only=True per ottenere valori calcolati
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    # Conta X nel foglio borderò (A11:A612)
    ws_bordero = wb['borderò']
    totale_eseguite = 0
    for row in range(11, 613):
        col_a = ws_bordero.cell(row=row, column=1).value
        if col_a and str(col_a).upper() == 'X':
            totale_eseguite += 1
    
    # Conta X nel foglio Display
    ws_display = wb['Display']
    display_eseguite = 0
    for row in ws_display.iter_rows():
        if row[0] and str(row[0]).upper() == 'X':
            display_eseguite += 1
    
    # Estrai Top 5 coreografie da colonne D (coreografia) e G (richieste)
    coreografie_dict = {}
    for row in range(12, 613):
        col_d = ws_bordero.cell(row=row, column=4).value  # Coreografia
        col_g = ws_bordero.cell(row=row, column=7).value  # Richieste
        
        if col_d and col_g:
            coreografia = str(col_d).strip()
            try:
                conteggio = int(col_g) if col_g else 0
            except:
                conteggio = 0
            
            if coreografia and conteggio > 0:
                if coreografia not in coreografie_dict:
                    coreografie_dict[coreografia] = 0
                coreografie_dict[coreografia] += conteggio
    
    sorted_coreografie = sorted(coreografie_dict.items(), key=lambda x: x[1], reverse=True)
    top_5 = sorted_coreografie[:5]
    
    # Calcola totale richieste
    totale_richieste = sum(count for _, count in sorted_coreografie)
    non_eseguite_richieste = totale_richieste - display_eseguite if totale_richieste > display_eseguite else 0
    
    # Crea il JSON
    data = {
        'totale_eseguite': totale_eseguite,
        'display_eseguite': display_eseguite,
        'coreografia_piu_richiesta': top_5[0][0] if top_5 else '',
        'conteggio_piu_richiesta': top_5[0][1] if top_5 else 0,
        'top_5': [{'nome': nome, 'conteggio': conteggio} for nome, conteggio in top_5],
        'eseguite': display_eseguite,
        'non_eseguite': non_eseguite_richieste,
        'totale_richieste': totale_richieste,
        'fonte': 'Borderò - ver 13.1.46.xlsm'
    }
    
    # Salva il JSON nella cartella public (per il deployment)
    output_path = r'c:\VSC_Live_Server\public\border_data.json'
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f'✓ File aggiornato: {output_path}')
    print(f'  - Eseguite: {totale_eseguite}')
    print(f'  - Coreografia più richiesta: {data["coreografia_piu_richiesta"]} ({data["conteggio_piu_richiesta"]} volte)')
    print(f'  - Top 5 aggiornate')
    
    return True

if __name__ == '__main__':
    try:
        update_border_stats()
    except Exception as e:
        print(f'Errore: {e}')
        import traceback
        traceback.print_exc()
